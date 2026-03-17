"""
Do Buybacks Explain European Stock Returns Better Than Earnings Growth?
=======================================================================
Regression analysis across FTSE MIB, DAX 40, CAC 40, IBEX 35 (2016-2024)

Requirements:
    pip install pandas numpy statsmodels openpyxl

Input:
    buyback_analysis_european_markets.xlsx (same directory)

Output:
    buyback_regression_results.xlsx
"""

import pandas as pd
import numpy as np
import statsmodels.api as sm
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings("ignore")

INPUT_FILE = "buyback_analysis_european_markets.xlsx"
OUTPUT_FILE = "buyback_regression_results.xlsx"
SEPARATOR_TICKERS = ["FTSE MIB", "DAX 40", "CAC 40", "IBEX 35"]
DUPLICATE_TICKERS = ["STLAP.PA", "STM.PA"]  # Stellantis + STMicro duplicates
WINSORIZE_BOUNDS = (5, 95)
PE_FLOOR = 0
PE_CEILING = 200


# ============================================================
# 1. LOAD DATA
# ============================================================

xls = pd.ExcelFile(INPUT_FILE)

# EPS Decomposition — already in long format (one row per company-year)
eps = pd.read_excel(xls, sheet_name="EPS Decomposition", header=2)
eps = eps[[
    "Ticker", "Company", "Country", "Index", "Year",
    "NI Growth (%)", "Share Count Change (%)", "EPS Growth (%)",
    "Income Effect (%)", "Buyback Effect (%)"
]].copy()

# Stock Price Return — wide format, needs melting
ret_wide = pd.read_excel(xls, sheet_name="Stock Price Return", header=2)
ret_wide = ret_wide[~ret_wide["Ticker"].isin(SEPARATOR_TICKERS)].copy()
year_cols = [c for c in ret_wide.columns if isinstance(c, (int, float)) or (isinstance(c, str) and c.isdigit())]
ret_long = ret_wide.melt(id_vars=["Ticker"], value_vars=year_cols, var_name="Year", value_name="Return")
ret_long["Year"] = ret_long["Year"].astype(int)

# PE Calculated — same wide layout
pe_wide = pd.read_excel(xls, sheet_name="PE Calculated", header=2)
pe_wide = pe_wide[~pe_wide["Ticker"].isin(SEPARATOR_TICKERS)].copy()
pe_long = pe_wide.melt(id_vars=["Ticker"], value_vars=year_cols, var_name="Year", value_name="PE")
pe_long["Year"] = pe_long["Year"].astype(int)


# ============================================================
# 2. MERGE AND CLEAN
# ============================================================

df = eps.merge(ret_long, on=["Ticker", "Year"], how="inner")
df = df.merge(pe_long, on=["Ticker", "Year"], how="inner")

# Remove duplicate cross-listed companies
df = df[~df["Ticker"].isin(DUPLICATE_TICKERS)].copy()

# Compute PE Change: need prior-year PE
pe_prior = pe_long.copy()
pe_prior["Year"] = pe_prior["Year"] + 1
pe_prior = pe_prior.rename(columns={"PE": "PE_prior"})
df = df.merge(pe_prior[["Ticker", "Year", "PE_prior"]], on=["Ticker", "Year"], how="left")
df["PE_Change"] = (df["PE"] - df["PE_prior"]) / df["PE_prior"].abs()

# Filter out unusable PE values
df = df[df["PE"].notna() & df["PE_prior"].notna()].copy()
df = df[
    (df["PE"] > PE_FLOOR) & (df["PE"] < PE_CEILING) &
    (df["PE_prior"] > PE_FLOOR) & (df["PE_prior"] < PE_CEILING)
].copy()

# Drop any remaining missing values
df = df.dropna(subset=[
    "Return", "NI Growth (%)", "EPS Growth (%)",
    "Share Count Change (%)", "PE_Change"
]).copy()

print(f"Final sample: {len(df)} company-year observations")
print(f"Index breakdown:\n{df['Index'].value_counts().to_string()}\n")


# ============================================================
# 3. WINSORIZE
# ============================================================

winsorize_cols = [
    "NI Growth (%)", "EPS Growth (%)", "Share Count Change (%)",
    "Return", "PE_Change"
]

for col in winsorize_cols:
    lo, hi = np.nanpercentile(df[col].values, WINSORIZE_BOUNDS)
    df[col + "_w"] = np.clip(df[col].values, lo, hi)


# ============================================================
# 4. REGRESSIONS
# ============================================================

def run_ols(y, X):
    """OLS with HC1 robust standard errors."""
    X = sm.add_constant(X)
    return sm.OLS(y, X).fit(cov_type="HC1")

def sig_stars(p):
    if p < 0.01: return "***"
    if p < 0.05: return "**"
    if p < 0.10: return "*"
    return ""

# Model specifications (all use winsorized variables)
MODELS = {
    "M1": ["NI Growth (%)_w"],
    "M2": ["EPS Growth (%)_w"],
    "M3": ["NI Growth (%)_w", "Share Count Change (%)_w"],
    "M4": ["NI Growth (%)_w", "PE_Change_w"],
    "M5": ["NI Growth (%)_w", "Share Count Change (%)_w", "PE_Change_w"],
    "M6": ["EPS Growth (%)_w", "PE_Change_w"],
}

# Samples
SAMPLES = {
    "Pooled": df,
    "FTSE MIB": df[df["Index"] == "FTSE MIB"],
    "DAX 40": df[df["Index"] == "DAX 40"],
    "CAC 40": df[df["Index"] == "CAC 40"],
    "IBEX 35": df[df["Index"] == "IBEX 35"],
}

# Run all 30 regressions
results = {}
for sample_name, sample_df in SAMPLES.items():
    for model_name, x_cols in MODELS.items():
        results[(sample_name, model_name)] = run_ols(
            sample_df["Return_w"], sample_df[x_cols]
        )


# ============================================================
# 5. PRINT RESULTS TO CONSOLE
# ============================================================

# --- Table 1: R² comparison ---
print("=" * 85)
print("TABLE 1: R² Comparison Across Models")
print("=" * 85)
header = f"{'Sample':<12} {'N':>5} {'M1':>8} {'M2':>8} {'M3':>8} {'M4':>8} {'M5':>8} {'M6':>8}"
print(header)
print("-" * 75)
for name, sub in SAMPLES.items():
    r2s = " ".join(f"{results[(name, m)].rsquared:>8.4f}" for m in MODELS)
    print(f"{name:<12} {len(sub):>5} {r2s}")

# --- Table 2: Model 5 coefficients ---
print(f"\n{'=' * 85}")
print("TABLE 2: Model 5 Coefficients (Return ~ NI Growth + Share Count Change + PE Change)")
print("=" * 85)
var_labels = {
    "const": "Intercept",
    "NI Growth (%)_w": "NI Growth (%)",
    "Share Count Change (%)_w": "Share Count Change (%)",
    "PE_Change_w": "PE Change",
}

for name in SAMPLES:
    m = results[(name, "M5")]
    print(f"\n--- {name} (N={len(SAMPLES[name])}, R²={m.rsquared:.4f}) ---")
    print(f"{'Variable':<28} {'Coef':>10} {'t-stat':>10} {'p-value':>10} {'Sig':>5}")
    print("-" * 68)
    for var in m.params.index:
        label = var_labels.get(var, var)
        print(f"{label:<28} {m.params[var]:>10.4f} {m.tvalues[var]:>10.3f} "
              f"{m.pvalues[var]:>10.4f} {sig_stars(m.pvalues[var]):>5}")

# --- Table 3: M1 vs M5 ---
print(f"\n{'=' * 85}")
print("TABLE 3: Model Comparison — M1 vs M5")
print("=" * 85)
print(f"{'Sample':<12} {'M1 R²':>8} {'M5 R²':>8} {'Δ R²':>8}")
print("-" * 40)
for name in SAMPLES:
    m1 = results[(name, "M1")].rsquared
    m5 = results[(name, "M5")].rsquared
    print(f"{name:<12} {m1:>8.4f} {m5:>8.4f} {m5 - m1:>8.4f}")

# --- Table 4: Buyback intensity ---
print(f"\n{'=' * 85}")
print("TABLE 4: Buyback Intensity by Index")
print("=" * 85)
print(f"{'Sample':<12} {'N':>5} {'Mean':>8} {'Median':>8} {'% Buyback':>10}")
print("-" * 48)
for name, sub in SAMPLES.items():
    sc = sub["Share Count Change (%)_w"]
    print(f"{name:<12} {len(sub):>5} {sc.mean():>8.3f} {sc.median():>8.3f} "
          f"{(sc < 0).mean() * 100:>9.1f}%")

# --- Table 5: EPS decomposition by index ---
# (Uses full EPS decomposition data, not just regression sample)
eps_clean = eps[~eps["Ticker"].isin(DUPLICATE_TICKERS)].copy()
eps_clean = eps_clean.dropna(subset=[
    "Income Effect (%)", "Buyback Effect (%)", "EPS Growth (%)"
])
for col in ["Income Effect (%)", "Buyback Effect (%)", "EPS Growth (%)"]:
    lo, hi = np.nanpercentile(eps_clean[col].values, WINSORIZE_BOUNDS)
    eps_clean[col + "_w"] = np.clip(eps_clean[col].values, lo, hi)

eps_samples = {
    "Pooled": eps_clean,
    "FTSE MIB": eps_clean[eps_clean["Index"] == "FTSE MIB"],
    "DAX 40": eps_clean[eps_clean["Index"] == "DAX 40"],
    "CAC 40": eps_clean[eps_clean["Index"] == "CAC 40"],
    "IBEX 35": eps_clean[eps_clean["Index"] == "IBEX 35"],
}

print(f"\n{'=' * 85}")
print("TABLE 5: EPS Growth Decomposition by Index")
print("=" * 85)
print(f"{'Sample':<12} {'N':>5} {'EPS Gr':>9} {'Inc Eff':>9} {'BB Eff':>9} {'BB Share':>9}")
print("-" * 58)
for name, sub in eps_samples.items():
    avg_eps = sub["EPS Growth (%)_w"].mean() * 100
    avg_inc = sub["Income Effect (%)_w"].mean() * 100
    avg_bb = sub["Buyback Effect (%)_w"].mean() * 100
    bb_share = avg_bb / avg_eps * 100 if abs(avg_eps) > 0.001 else 0
    print(f"{name:<12} {len(sub):>5} {avg_eps:>8.2f}% {avg_inc:>8.2f}% {avg_bb:>8.2f}% "
          f"{bb_share:>8.1f}%")

# --- Table 6: EPS decomposition by year ---
print(f"\n{'=' * 85}")
print("TABLE 6: EPS Growth Decomposition by Year (Pooled)")
print("=" * 85)
print(f"{'Year':>6} {'N':>5} {'EPS Gr':>9} {'Inc Eff':>9} {'BB Eff':>9} {'BB Share':>9}")
print("-" * 50)
for year in sorted(eps_clean["Year"].unique()):
    sub = eps_clean[eps_clean["Year"] == year]
    avg_eps = sub["EPS Growth (%)_w"].mean() * 100
    avg_inc = sub["Income Effect (%)_w"].mean() * 100
    avg_bb = sub["Buyback Effect (%)_w"].mean() * 100
    bb_share = avg_bb / avg_eps * 100 if abs(avg_eps) > 0.001 else 0
    print(f"{year:>6} {len(sub):>5} {avg_eps:>8.2f}% {avg_inc:>8.2f}% {avg_bb:>8.2f}% "
          f"{bb_share:>8.1f}%")


# ============================================================
# 6. EXPORT TO EXCEL
# ============================================================

wb = Workbook()
HFONT = Font(name="Arial", bold=True, size=11)
TFONT = Font(name="Arial", size=10)
HFILL = PatternFill("solid", fgColor="D9E1F2")
PCT = "0.00%"

def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HFONT
        cell.fill = HFILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

def auto_width(ws):
    for col in ws.columns:
        mx = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(mx + 3, 22)


# Sheet 1: R² Comparison
ws = wb.active
ws.title = "R² Comparison"
ws["A1"] = "Table 1: R² Comparison Across Models"
ws["A1"].font = Font(name="Arial", bold=True, size=13)
ws["A2"] = "Dependent variable: winsorized annual stock price return. OLS with HC1 robust SEs."
ws["A2"].font = Font(name="Arial", italic=True, size=9, color="666666")

for i, h in enumerate(["Sample", "N", "M1", "M2", "M3", "M4", "M5", "M6"], 1):
    ws.cell(row=4, column=i, value=h)
style_header(ws, 4, 8)

for r, name in enumerate(SAMPLES, 5):
    ws.cell(row=r, column=1, value=name).font = TFONT
    ws.cell(row=r, column=2, value=len(SAMPLES[name])).font = TFONT
    for ci, mn in enumerate(MODELS, 3):
        c = ws.cell(row=r, column=ci, value=round(results[(name, mn)].rsquared, 4))
        c.font = TFONT
        c.number_format = "0.0000"
        c.alignment = Alignment(horizontal="center")

ws.cell(row=11, column=1, value="Model Specifications:").font = Font(name="Arial", bold=True, size=10)
for i, desc in enumerate([
    "M1: Return ~ NI Growth",
    "M2: Return ~ EPS Growth",
    "M3: Return ~ NI Growth + Share Count Change",
    "M4: Return ~ NI Growth + PE Change",
    "M5: Return ~ NI Growth + Share Count Change + PE Change",
    "M6: Return ~ EPS Growth + PE Change",
]):
    ws.cell(row=12 + i, column=1, value=desc).font = Font(name="Arial", size=9, color="444444")
auto_width(ws)


# Sheet 2: M5 Coefficients
ws2 = wb.create_sheet("M5 Coefficients")
ws2["A1"] = "Table 2: Model 5 — Return ~ NI Growth + Share Count Change + PE Change"
ws2["A1"].font = Font(name="Arial", bold=True, size=13)
ws2["A2"] = "OLS with HC1 robust SEs. Significance: *** p<0.01, ** p<0.05, * p<0.10"
ws2["A2"].font = Font(name="Arial", italic=True, size=9, color="666666")

row = 4
for name in SAMPLES:
    m = results[(name, "M5")]
    ws2.cell(row=row, column=1, value=f"{name} (N={len(SAMPLES[name])})").font = HFONT
    ws2.cell(row=row, column=6, value=f"R² = {m.rsquared:.4f}").font = Font(
        name="Arial", bold=True, size=10, color="1F4E79"
    )
    row += 1
    for ci, h in enumerate(["Variable", "Coef", "t-stat", "p-value", "Sig"], 1):
        ws2.cell(row=row, column=ci, value=h)
    style_header(ws2, row, 5)
    row += 1
    for var in m.params.index:
        ws2.cell(row=row, column=1, value=var_labels.get(var, var)).font = TFONT
        ws2.cell(row=row, column=2, value=round(m.params[var], 4)).font = TFONT
        ws2.cell(row=row, column=3, value=round(m.tvalues[var], 3)).font = TFONT
        ws2.cell(row=row, column=4, value=round(m.pvalues[var], 4)).font = TFONT
        s = sig_stars(m.pvalues[var])
        ws2.cell(row=row, column=5, value=s).font = Font(name="Arial", size=10, bold=bool(s))
        row += 1
    row += 1
auto_width(ws2)


# Sheet 3: M1 vs M5
ws3 = wb.create_sheet("M1 vs M5")
ws3["A1"] = "Table 3: Model Comparison — M1 vs M5"
ws3["A1"].font = Font(name="Arial", bold=True, size=13)

for ci, h in enumerate(["Sample", "M1 R²", "M5 R²", "Δ R²"], 1):
    ws3.cell(row=3, column=ci, value=h)
style_header(ws3, 3, 4)

for r, name in enumerate(SAMPLES, 4):
    m1r = results[(name, "M1")].rsquared
    m5r = results[(name, "M5")].rsquared
    ws3.cell(row=r, column=1, value=name).font = TFONT
    ws3.cell(row=r, column=2, value=round(m1r, 4)).font = TFONT
    ws3.cell(row=r, column=3, value=round(m5r, 4)).font = TFONT
    ws3.cell(row=r, column=4, value=round(m5r - m1r, 4)).font = Font(
        name="Arial", size=10, bold=True, color="1F4E79"
    )
auto_width(ws3)


# Sheet 4: Buyback Intensity
ws4 = wb.create_sheet("Buyback Intensity")
ws4["A1"] = "Table 4: Buyback Intensity by Index"
ws4["A1"].font = Font(name="Arial", bold=True, size=13)

for ci, h in enumerate(["Sample", "N", "Mean (%)", "Median (%)", "% Years with Buybacks"], 1):
    ws4.cell(row=3, column=ci, value=h)
style_header(ws4, 3, 5)

for r, name in enumerate(SAMPLES, 4):
    sc = SAMPLES[name]["Share Count Change (%)_w"]
    ws4.cell(row=r, column=1, value=name).font = TFONT
    ws4.cell(row=r, column=2, value=len(SAMPLES[name])).font = TFONT
    ws4.cell(row=r, column=3, value=round(sc.mean() * 100, 2)).font = TFONT
    ws4.cell(row=r, column=4, value=round(sc.median() * 100, 2)).font = TFONT
    ws4.cell(row=r, column=5, value=f"{(sc < 0).mean() * 100:.1f}%").font = TFONT
auto_width(ws4)


# Sheet 5: EPS Decomposition by Index
ws5 = wb.create_sheet("EPS Decomposition")
ws5["A1"] = "Table 5: EPS Growth Decomposition by Index"
ws5["A1"].font = Font(name="Arial", bold=True, size=13)

for ci, h in enumerate(["Sample", "N", "Avg EPS Growth", "Income Effect", "Buyback Effect",
                         "BB Share of EPS", "% Years BB > 0"], 1):
    ws5.cell(row=3, column=ci, value=h)
style_header(ws5, 3, 7)

for r, (name, sub) in enumerate(eps_samples.items(), 4):
    avg_eps = sub["EPS Growth (%)_w"].mean()
    avg_inc = sub["Income Effect (%)_w"].mean()
    avg_bb = sub["Buyback Effect (%)_w"].mean()
    bb_share = avg_bb / avg_eps if abs(avg_eps) > 0.00001 else 0
    pct_pos = (sub["Buyback Effect (%)_w"] > 0).mean()
    ws5.cell(row=r, column=1, value=name).font = TFONT
    ws5.cell(row=r, column=2, value=len(sub)).font = TFONT
    for ci, val in [(3, avg_eps), (4, avg_inc), (5, avg_bb), (6, bb_share), (7, pct_pos)]:
        c = ws5.cell(row=r, column=ci, value=val)
        c.font = TFONT
        c.number_format = PCT
        c.alignment = Alignment(horizontal="center")
auto_width(ws5)


# Sheet 6: EPS Decomposition by Year
ws6 = wb.create_sheet("Decomposition by Year")
ws6["A1"] = "Table 6: EPS Growth Decomposition by Year (Pooled)"
ws6["A1"].font = Font(name="Arial", bold=True, size=13)

for ci, h in enumerate(["Year", "N", "Avg EPS Growth", "Income Effect", "Buyback Effect",
                         "BB Share of EPS"], 1):
    ws6.cell(row=3, column=ci, value=h)
style_header(ws6, 3, 6)

for r, year in enumerate(sorted(eps_clean["Year"].unique()), 4):
    sub = eps_clean[eps_clean["Year"] == year]
    avg_eps = sub["EPS Growth (%)_w"].mean()
    avg_inc = sub["Income Effect (%)_w"].mean()
    avg_bb = sub["Buyback Effect (%)_w"].mean()
    bb_share = avg_bb / avg_eps if abs(avg_eps) > 0.00001 else 0
    ws6.cell(row=r, column=1, value=int(year)).font = TFONT
    ws6.cell(row=r, column=2, value=len(sub)).font = TFONT
    for ci, val in [(3, avg_eps), (4, avg_inc), (5, avg_bb), (6, bb_share)]:
        c = ws6.cell(row=r, column=ci, value=val)
        c.font = TFONT
        c.number_format = PCT
        c.alignment = Alignment(horizontal="center")
auto_width(ws6)


wb.save(OUTPUT_FILE)
print(f"\nResults saved to {OUTPUT_FILE}")
